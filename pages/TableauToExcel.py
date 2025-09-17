""" 
# Purpose: This python script integrates multiple excel files into one excel file. Script -> Front-end Development
"""

# import streamlit as st
# import openpyxl
# from copy import copy
# import tempfile
# import os

# st.title("Tableau -> Report Web Application")

# uploaded_files = st.file_uploader(
#     "Upload Excel files",
#     type = "xlsx",
#     accept_multiple_files = True
# )

# def copy_sheet(source_sheet, target_wb, new_title):
#     target_sheet = target_wb.create_sheet(title=new_title[:31])

#     # Copy cell values and styles
#     for row in source_sheet.iter_rows():
#         for cell in row:
#             new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
#             if cell.has_style:
#                 new_cell.font = copy(cell.font)
#                 new_cell.border = copy(cell.border)
#                 new_cell.fill = copy(cell.fill)
#                 new_cell.number_format = cell.number_format
#                 new_cell.protection = copy(cell.protection)
#                 new_cell.alignment = copy(cell.alignment)

#     # Copy column widths
#     for col_letter, col_dim in source_sheet.column_dimensions.items():
#         target_sheet.column_dimensions[col_letter].width = col_dim.width

#     # Copy row heights
#     for row_idx, row_dim in source_sheet.row_dimensions.items():
#         target_sheet.row_dimensions[row_idx].height = row_dim.height

#     # Copy merged cells
#     for merged_range in source_sheet.merged_cells.ranges:
#         target_sheet.merge_cells(str(merged_range))

#     return target_sheet

# if uploaded_files:
#     temp_dir = tempfile.mkdtemp()
#     file_paths = []

#     # Save uploaded files to temp folder
#     for f in uploaded_files:
#         path = os.path.join(temp_dir, f.name)
#         with open(path, "wb") as temp_file:
#             temp_file.write(f.read())
#         file_paths.append(path)

#     # Create combined workbook
#     combined_wb = openpyxl.Workbook()
#     combined_wb.remove(combined_wb.active)  # Remove default sheet

#     for file_path in file_paths:
#         wb = openpyxl.load_workbook(file_path)
#         for sheet in wb.worksheets:
#             copy_sheet(sheet, combined_wb, sheet.title)
#         wb.close()

#     output_file = os.path.join(temp_dir, "Combined.xlsx")
#     combined_wb.save(output_file)

#     # Provide download button
#     with open(output_file, "rb") as f:
#         st.download_button(
#             "Download Combined Excel",
#             data=f,
#             file_name="Combined.xlsx"
#         )

import streamlit as st
import openpyxl
from copy import copy
import tempfile
import os

st.title("Excel Combiner (Preserve Formatting)")

uploaded_files = st.file_uploader(
    "Upload Excel files",
    type="xlsx",
    accept_multiple_files=True
)

def copy_sheet(source_sheet, target_wb, new_title):
    # Ensure sheet name is <=31 chars
    new_title = new_title[:31]

    # Handle duplicate names
    existing_titles = [s.title for s in target_wb.worksheets]
    original_title = new_title
    counter = 1
    while new_title in existing_titles:
        new_title = f"{original_title}_{counter}"[:31]
        counter += 1

    target_sheet = target_wb.create_sheet(title=new_title)

    # Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copy column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width

    # Copy row heights
    for row_idx, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = row_dim.height

    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    return target_sheet

if uploaded_files:
    temp_dir = tempfile.mkdtemp()
    file_paths = []

    # Save uploaded files to temp folder
    for f in uploaded_files:
        path = os.path.join(temp_dir, f.name)
        with open(path, "wb") as temp_file:
            temp_file.write(f.read())
        file_paths.append(path)

    # Create combined workbook
    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)  # Remove default sheet

    for file_path in file_paths:
        wb = openpyxl.load_workbook(file_path)
        # Use the file name (without extension) as the sheet name
        sheet_name = os.path.splitext(os.path.basename(file_path))[0]
        # Copy only the first sheet of the file (or you can choose any)
        copy_sheet(wb.worksheets[0], combined_wb, sheet_name)
        wb.close()

    output_file = os.path.join(temp_dir, "Combined.xlsx")
    combined_wb.save(output_file)

    # Provide download button
    with open(output_file, "rb") as f:
        st.download_button(
            "Download Combined Excel",
            data=f,
            file_name="Combined.xlsx"
        )
